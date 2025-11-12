from flask import Flask, request, render_template_string, redirect
from datetime import datetime

app = Flask(__name__)
gastos = []
saldo_inicial = 0.0

# ⚠️ Este backend está desativado no momento.
# O app usa LocalStorage via JavaScript para armazenar dados localmente.
# Este código será útil para implementação de login e banco de dados no futuro.

@app.route("/", methods=["GET", "POST"])
def index():
    global saldo_inicial

    if request.method == "POST":
        if "editar_saldo" in request.form:
            try:
                saldo_inicial = float(request.form["editar_saldo"].replace(",", "."))
            except ValueError:
                return "<h3>❌ Saldo inválido. Use números com vírgula ou ponto.</h3><a href='/'>Voltar</a>"
        elif "saldo" in request.form:
            try:
                saldo_inicial = float(request.form["saldo"].replace(",", "."))
            except ValueError:
                return "<h3>❌ Saldo inválido. Use números com vírgula ou ponto.</h3><a href='/'>Voltar</a>"
        else:
            try:
                valor = float(request.form["valor"].replace(",", "."))
            except ValueError:
                return "<h3>❌ Valor inválido. Use números com vírgula ou ponto.</h3><a href='/'>Voltar</a>"

            categoria = request.form["categoria"]
            data = request.form["data"]
            data_formatada = f"{data[:2]}/{data[2:4]}/{data[4:]}"
            gasto = {
                "valor": valor,
                "categoria": categoria,
                "data": data_formatada,
            }
            gastos.append(gasto)

    total = sum(g["valor"] for g in gastos)
    saldo_restante = saldo_inicial - total if saldo_inicial is not None else 0.0
    

    html = """
    <div class='cabecalho'>
        <h1>Controle de Gastos Pessoais</h1>
    </div>  
    <head>
        <link rel="stylesheet" href="/static/style.css">
        <script src="https://cdn.sheetjs.com/xlsx-latest/package/dist/xlsx.full.min.js"></script>
        <script src="/static/script.js" defer></script>
    </head>

    <h2>Definir Saldo Inicial</h2>
    <label for="saldo-inicial">Digite quanto você tem disponível (R$):</label>
    <input type="text" id="saldo-inicial" placeholder="Ex: 1000,00">
    <button onclick="definirSaldo()">Definir ou Editar saldo</button>

    <p><strong>Saldo disponível:</strong> R$ <span id="saldo-disponivel">0.00</span></p>

    <h2>Adicionar Gasto</h2>
    <label for="valor">Digite o valor do gasto (R$):</label>
    <input type="text" id="valor"><br>
    <label for="categoria">Digite a categoria (ex: mercado, transporte):</label>
    <input type="text" id="categoria"><br>
    <label for="data">Digite a data (DDMMAAAA):</label>
    <input type="text" id="data"><br>
    <button onclick="adicionarGasto()">Adicionar gasto</button>

    <h2>Tabela de Gastos</h2>
    <div class="tabela-container">
        <div class="tabela-scroll">
            <table id="tabela-gastos">
                <tr>
                    <th>Valor</th>
                    <th>Categoria</th>
                    <th>Data</th>
                    <th>Ações</th>
                </tr>
            </table>
        </div>
    </div>

    <p><strong>Total de gastos:</strong> R$ <span id="total-gastos">0.00</span></p>
    <p><strong>Saldo inicial:</strong> <span id="saldo-inicial-final">0.00</span></p>
    <p><strong>Saldo restante:</strong> <span id="saldo-restante">0.00</span></p>

    <button onclick="exportarCSV()">
        <img src="/static/icons/csv.png" style="width:20px; vertical-align:middle; margin-right:6px;">
        Exportar CSV
    </button>

    <button onclick="exportarExcel()">
        <img src="/static/icons/excel.png" style="width:20px; vertical-align:middle; margin-right:6px;">
        Exportar Excel
    </button>
    """
    return render_template_string(html, gastos=gastos, total=total, saldo_inicial=saldo_inicial, saldo=saldo_restante)
    

@app.route("/excluir", methods=["POST"])
def excluir():
    indice = int(request.form["indice"])
    if 0 <= indice < len(gastos):
        gastos.pop(indice)
    return redirect("/")

import csv
from io import StringIO
from flask import Response

@app.route("/exportar", methods=["GET"])
def exportar():
    agora = datetime.now()
    mes_ano = agora.strftime("%B %Y")  # Ex: "October 2025"

    output = StringIO()
    writer = csv.writer(output)

    # Cabeçalho com mês e ano
    writer.writerow([f"Relatório de Gastos - {mes_ano}"])
    writer.writerow([])

    # Saldo inicial, total e restante
    total = sum(g["valor"] for g in gastos)
    saldo_restante = saldo_inicial - total if saldo_inicial is not None else 0.0

    writer.writerow(["Saldo Inicial", saldo_inicial if saldo_inicial is not None else 0.0])
    writer.writerow(["Total de Gastos", total])
    writer.writerow(["Saldo Restante", saldo_restante])
    writer.writerow([])

    # Tabela de gastos
    writer.writerow(["Valor", "Categoria", "Data"])
    for g in gastos:
        writer.writerow([g["valor"], g["categoria"], g["data"]])

    output.seek(0)
    filename = f"gastos_{agora.strftime('%m_%Y')}.csv"  # Ex: gastos_10_2025.csv

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import send_file
from io import BytesIO

@app.route("/exportar_excel", methods=["GET"])
def exportar_excel():
    agora = datetime.now()
    mes_ano = agora.strftime("%B %Y")  # Ex: "Outubro 2025"

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos do Mês"

    # Cabeçalho
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Relatório de Gastos - {mes_ano}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Dados financeiros
    total = sum(g["valor"] for g in gastos)
    saldo_restante = saldo_inicial - total if saldo_inicial is not None else 0.0

    ws.append([])
    ws.append(["Saldo Inicial", saldo_inicial if saldo_inicial is not None else 0.0])
    ws.append(["Total de Gastos", total])
    ws.append(["Saldo Restante", saldo_restante])
    ws.append([])

    # Tabela de gastos
    ws.append(["Valor", "Categoria", "Data"])
    for g in gastos:
        ws.append([g["valor"], g["categoria"], g["data"]])

    # Estilizar cabeçalhos
    for cell in ws[ws.max_row - len(gastos) - 1]:
        cell.font = Font(bold=True)

    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"gastos_{agora.strftime('%m_%Y')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=True)

